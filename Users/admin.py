from django.contrib import admin
from django.contrib.auth.admin import UserAdmin
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter
from .forms import (
    CustomAdminPasswordChangeForm,
    CustomUserAdminChangeForm,
    CustomUserAdminCreationForm,
)
from .models import AuditTrail, CustomUser, PasswordHistory, RoleModuleAccess, SystemModule, SystemSetting, UserModuleAccess

class CustomUserAdmin(UserAdmin):
    form = CustomUserAdminChangeForm
    add_form = CustomUserAdminCreationForm
    change_password_form = CustomAdminPasswordChangeForm
    readonly_fields = ("last_login", "date_joined", "password_changed_at")

    fieldsets = (
        (None, {"fields": ("email", "password")}),
        (
            "Personal info",
            {"fields": ("name", "surname", "phone_number", "address", "department", "gender")},
        ),
        (
            "Security",
            {
                "fields": (
                    "must_change_password",
                    "password_changed_at",
                    "microsoft_authenticator_enabled",
                    "microsoft_authenticator_confirmed_at",
                )
            },
        ),
        (
            "Permissions",
            {"fields": ("is_active", "is_staff", "is_superuser", "groups", "user_permissions")},
        ),
        ("Important dates", {"fields": ("last_login", "date_joined")}),
    )

    add_fieldsets = (
        (
            None,
            {
                "classes": ("wide",),
                "fields": (
                    "email",
                    "name",
                    "surname",
                    "phone_number",
                    "address",
                    "department",
                    "gender",
                    "password1",
                    "password2",
                    "is_active",
                    "is_staff",
                    "is_superuser",
                    "groups",
                    "user_permissions",
                ),
            },
        ),
    )

    list_display = (
        "email",
        "name",
        "surname",
        "must_change_password",
        "password_changed_at",
        "is_staff",
        "is_active",
    )
    list_filter = ("is_active", "is_staff", "is_superuser", "must_change_password")
    search_fields = ("email", "name", "surname")
    ordering = ("email",)

admin.site.register(CustomUser, CustomUserAdmin)


@admin.register(AuditTrail)
class AuditTrailAdmin(admin.ModelAdmin):
    list_display = ('user__email', 'model_name', 'action', 'object_id', 'timestamp')
    list_filter = ('model_name', 'action', 'timestamp')
    search_fields = ('user__email', 'model_name', 'object_id', 'change_description')
    ordering = ('-timestamp',)

    # Add custom actions
    actions = ['export_as_excel']

    def export_as_excel(self, request, queryset):
        """
        Export selected audit trail records to an Excel file.
        """
        # Create a workbook and sheet
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Audit Trails"

        # Define column headers
        headers = ['User Email', 'Model Name', 'Action', 'Object ID', 'Timestamp', 'Change Description']
        sheet.append(headers)

        # Add data rows
        for record in queryset:
            row = [
                record.user.email if record.user else 'N/A',
                record.model_name,
                record.action,
                record.object_id,
                record.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                record.change_description,
            ]
            sheet.append(row)

        # Adjust column widths
        for col_num, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            sheet.column_dimensions[col_letter].width = 25

        # Create HTTP response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="AuditTrails.xlsx"'
        workbook.save(response)

        return response

    export_as_excel.short_description = "Export selected records to Excel"


@admin.register(SystemModule)
class SystemModuleAdmin(admin.ModelAdmin):
    list_display = ("code", "name", "route_name", "is_active", "display_order")
    list_filter = ("is_active",)
    search_fields = ("code", "name", "route_name")
    ordering = ("display_order", "name")


@admin.register(RoleModuleAccess)
class RoleModuleAccessAdmin(admin.ModelAdmin):
    list_display = ("group", "module", "can_view", "updated_at")
    list_filter = ("can_view", "module")
    search_fields = ("group__name", "module__name")
    ordering = ("group__name", "module__display_order")


@admin.register(UserModuleAccess)
class UserModuleAccessAdmin(admin.ModelAdmin):
    list_display = ("user", "module", "can_view", "updated_at")
    list_filter = ("can_view", "module")
    search_fields = ("user__email", "module__name")
    ordering = ("user__email", "module__display_order")


@admin.register(SystemSetting)
class SystemSettingAdmin(admin.ModelAdmin):
    list_display = (
        "idle_timeout_minutes",
        "absolute_session_timeout_minutes",
        "default_landing_rule",
        "failed_login_limit",
        "lockout_duration_minutes",
        "enable_self_profile_edit",
        "enable_self_password_change",
        "password_expiry_days",
        "password_history_count",
        "password_policy",
        "enable_microsoft_authentication",
        "updated_at",
    )
    readonly_fields = ("updated_at",)


@admin.register(PasswordHistory)
class PasswordHistoryAdmin(admin.ModelAdmin):
    list_display = ("user", "created_at")
    search_fields = ("user__email", "user__name", "user__surname")
    ordering = ("-created_at",)
    readonly_fields = ("user", "password_hash", "created_at")

    def has_add_permission(self, request):
        return False
